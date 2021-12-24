

import openpyxl
wb = openpyxl.load_workbook("gestion des comptes.xlsx", data_only=True)

sheet = wb["Feuil1"]
print(sheet.cell(18,17).value)

som = 0
for x in range(1,sheet.max_row):
    type_donnee = sheet.cell(x,17).value
    # if sheet.cell(x,17).value == "Alimentation":
    #     print(str(sheet.cell(x,17).value )+ " : "+ str(int(sheet.cell(x,18).value) * -1)+"€")
    #     som+= int(sheet.cell(x,18).value)

    match type_donnee :
        case "Alimentation":
            print(str(sheet.cell(x,17).value )+ " : "+ str(int(sheet.cell(x,18).value) * -1)+"€")
            som+= int(sheet.cell(x,18).value)
        case "Retrait":
            print(str(sheet.cell(x,17).value )+ " : "+ str(int(sheet.cell(x,18).value) * -1)+"€")
            som+= int(sheet.cell(x,18).value)
        case "Shooping ":
            print(str(sheet.cell(x,17).value )+ " : "+ str(int(sheet.cell(x,18).value) * -1)+"€")
            som+= int(sheet.cell(x,18).value)
        case "Sport":
            print(str(sheet.cell(x,17).value )+ " : "+ str(int(sheet.cell(x,18).value) * -1)+"€")
            som+= int(sheet.cell(x,18).value)
        case "Telecom":
            print(str(sheet.cell(x,17).value )+ " : "+ str(int(sheet.cell(x,18).value) * -1)+"€")
            som+= int(sheet.cell(x,18).value)
        case "Assurance":
            print(str(sheet.cell(x,17).value )+ " : "+ str(int(sheet.cell(x,18).value) * -1)+"€")
            som+= int(sheet.cell(x,18).value)
        case "Energie":
            print(str(sheet.cell(x,17).value )+ " : "+ str(int(sheet.cell(x,18).value) * -1)+"€")
            som+= int(sheet.cell(x,18).value)
        case "Auto":
            print(str(sheet.cell(x,17).value )+ " : "+ str(int(sheet.cell(x,18).value) * -1)+"€")
            som+= int(sheet.cell(x,18).value)
        case "Loyer":
            print(str(sheet.cell(x,17).value )+ " : "+ str(int(sheet.cell(x,18).value) * -1)+"€")
            som+= int(sheet.cell(x,18).value)
        case "Autres":
            print(str(sheet.cell(x,17).value )+ " : "+ str(int(sheet.cell(x,18).value) * -1)+"€")
            som+= int(sheet.cell(x,18).value)
        case "Loisirs":
            print(str(sheet.cell(x,17).value )+ " : "+ str(int(sheet.cell(x,18).value) * -1)+"€")
            som+= int(sheet.cell(x,18).value)
print("")


print("Le total pour l'alimentation et Retrait est : "+ str(som * -1)+"€")