"Extraire des données depuis un fichier excel : "

import openpyxl

wb1 = openpyxl.load_workbook("novembre.xlsx", data_only=True)
wb2 = openpyxl.load_workbook("octobre.xlsx", data_only=True)
wb3 = openpyxl.load_workbook("decembre.xlsx", data_only=True)
"le nom de la feuille"
# print(wb.sheetnames)

"Feuille = sheet"
"Quand tu veux une feuille spécifique du fichier excel"
# sheet = wb["Feuil1"]
"Quand tu veux juste accéder à la feuille active"
sheet = wb1.active
sheet2 = wb2.active
sheet3 = wb3.active
sheet.max_column
"Recuperer des données dans le fichier avec maxcolumn & maxrow"
# for row in range(2,sheet.max_row+1):
#     for column in range(1,sheet.max_column+1):
#     # cell1 = sheet.cell(3,3)
#         print(sheet.cell(row,column).value, end = "  ")
#     print("-----")

"Recuperation des données que des champs non nul"

# "NOVEMBRE"

# mon_dico_fruit = {}
# for row in range(2,sheet.max_row+1):
#     fruit= sheet.cell(row,1).value
#     total_vente = sheet.cell(row,4).value
#     if not fruit:
#         break
#     "!!!AJOUT DES DONNEES DANS LE DICO!!!"
#     mon_dico_fruit[fruit] = total_vente
#     # print(fruit)
"Recup en fonction des données dans les 3 fichiers excels"
# mon_dico_oct = {}
# mon_dico_nov = {}
# mon_dico_dec = {}
# def recup_fichier(feuille,dico):
#     for row in range(2,feuille.max_row+1):
#         fruit= feuille.cell(row,1).value
#         total_vente = feuille.cell(row,4).value
#         if not fruit:
#             break
#         "!!!AJOUT DES DONNEES DANS LE DICO!!!"
#         dico[fruit] = total_vente
#     return dico
#     # print(fruit)

# print("Pour le mois d'Oct : ",recup_fichier(sheet2,mon_dico_oct))
# print("Pour le mois d'Nov : ",recup_fichier(sheet,mon_dico_nov))
# print("Pour le mois d'Dec : ",recup_fichier(sheet3,mon_dico_dec))

mon_dico_fruit = {}
"utiliser une fonction pour récupérer toutes les données des tab"
def recup_donnees(fichier,dic):
    for row in range(2,fichier.max_row+1):
        fruit= fichier.cell(row,1).value
        total_vente = fichier.cell(row,4).value
        if not fruit:
            break
        "!!!AJOUT DES DONNEES DANS LE DICO!!! --- si il est vide crée un tab, si il esxite ajoute une value supplementaire---"
        if dic.get(fruit):
            dic[fruit].append(total_vente)
        else:
            dic[fruit] = [total_vente]
    # print(fruit)
    return 

recup_donnees(sheet,mon_dico_fruit)
recup_donnees(sheet2,mon_dico_fruit)
recup_donnees(sheet3,mon_dico_fruit)


# print(mon_dico_fruit.keys())
# print(mon_dico_fruit)

# "Cette opération est plus long pour faire la somme totale de vente, voir ci-dessous "
# for fruit in mon_dico_fruit:
#     som = 0
#     for i in range(len(mon_dico_fruit[fruit])):
#         som +=mon_dico_fruit[fruit][i]
#     mon_dico_fruit[fruit].append(som)
# print(mon_dico_fruit)




"Faire une somme total de tab dans le dico"
for k in mon_dico_fruit.keys():
    # print(mon_dico_fruit[k])
    tatal_vente = sum(mon_dico_fruit[k])
    mon_dico_fruit[k].append(tatal_vente)

# print(mon_dico_fruit)
"CREE UNE FEUILLE DE SORTIE"
# wb_sortie = openpyxl.Workbook()
# sheet = wb_sortie.active
# sheet["A1"]="Article"
# sheet["B1"]="Octobre"
# sheet["C1"]="Novembre"
# sheet["D1"]="Decembre"
# sheet["E1"]="Total"
# ligne = 2
# "Remplissage de la nouvelle feuille de sortie, !!!Recuperer des données dans un dico!!!"
# for k in mon_dico_fruit.items():
#     nom_fruit = k[0]
#     ventes_fruit = k[1]
#     sheet.cell(ligne,1).value=nom_fruit
#     for j in range(0,len(ventes_fruit)):
#         sheet.cell(ligne,j+2).value=ventes_fruit[j]
#         # sheet.cell(ligne,3).value=k[1][1]
#         # sheet.cell(ligne,4).value=k[1][2]
#         # sheet.cell(ligne,5).value=k[1][3]
#     ligne += 1
#     # print(k)
# wb_sortie.save("Total_vente.xlsx")



# "Creation d'une nouvelle feuille dans un fichier excel"

# sheet_sortie = wb1.create_sheet("Total")
# sheet_sortie["A1"]="Article"
# sheet_sortie["B1"]="Octobre"
# sheet_sortie["C1"]="Novembre"
# sheet_sortie["D1"]="Decembre"
# sheet_sortie["E1"]="Total"
# ligne = 2
# "Remplissage de la nouvelle feuille de sortie, !!!Recuperer des données dans un dico!!!"
# for k in mon_dico_fruit.items():
#     nom_fruit = k[0]
#     ventes_fruit = k[1]
#     sheet_sortie.cell(ligne,1).value=nom_fruit
#     for j in range(0,len(ventes_fruit)):
#         sheet_sortie.cell(ligne,j+2).value=ventes_fruit[j]
#         # sheet.cell(ligne,3).value=k[1][1]
#         # sheet.cell(ligne,4).value=k[1][2]
#         # sheet.cell(ligne,5).value=k[1][3]
#     ligne += 1